import os
import math
import subprocess
import sys
import uuid
import json
import hmac
import base64
import hashlib
from collections import defaultdict
from datetime import date, datetime, time as dt_time
from urllib.parse import urlencode
from urllib import error as url_error
from urllib.request import Request, urlopen

from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv
from flask import Flask, jsonify, redirect, render_template, request, session, url_for
from flask_mysqldb import MySQL

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled by runtime message
    load_workbook = None

load_dotenv(".env")
load_dotenv(".env.local", override=True)

app = Flask(__name__)
app.secret_key = "secret123"
app.config["STATUS_UPDATE_TOKEN"] = os.getenv("STATUS_UPDATE_TOKEN", app.secret_key)

app.config["MYSQL_HOST"] = "208.91.199.36"
app.config["MYSQL_USER"] = "igssewjk_Iiot"
app.config["MYSQL_PASSWORD"] = "Sgi@admin"
app.config["MYSQL_DB"] = "igssewjk_Gggg"

mysql = MySQL(app)
scheduler = BackgroundScheduler()

ALLOWED_UPLOAD_EXTENSIONS = {".xlsx", ".csv"}
PHONE_HEADERS = {"phone", "phone_number", "mobile", "number", "contact", "contact_number"}
TIME_HEADERS = {"time", "schedule_time", "scheduled_at", "call_time", "datetime"}
NAME_HEADERS = {"name", "customer_name", "contact_name", "full_name"}
CALL_RECORD_ACCOUNT_ID = os.getenv("CALL_RECORD_ACCOUNT_ID", "")
CALL_RECORD_AUTH_ID = os.getenv("CALL_RECORD_AUTH_ID", CALL_RECORD_ACCOUNT_ID)
CALL_RECORD_AUTH_TOKEN = os.getenv("CALL_RECORD_AUTH_TOKEN", "")
CALL_RECORD_API_URL = os.getenv("CALL_RECORD_API_URL", "")
SELL_RATE_PER_MINUTE = float(os.getenv("SELL_RATE_PER_MINUTE", "5"))
WALLET_RECHARGE_URL = os.getenv("WALLET_RECHARGE_URL", "https://rzp.io/rzp/jey2OJZ")
MIN_RECHARGE_AMOUNT = float(os.getenv("MIN_RECHARGE_AMOUNT", "30"))
MAX_RECHARGE_AMOUNT = float(os.getenv("MAX_RECHARGE_AMOUNT", "10000"))
RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID", "")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET", "")
RAZORPAY_API_URL = os.getenv("RAZORPAY_API_URL", "https://api.razorpay.com/v1")
DEFAULT_ADMIN_EMAIL = os.getenv("DEFAULT_ADMIN_EMAIL", "pranjalsingh20032007@gmail.com").strip().lower()
SCHEDULER_ENABLED = os.getenv("SCHEDULER_ENABLED", "1") == "1"
_scheduler_bootstrapped = False


def login_required():
    if "user" not in session:
        return False
    return True


def admin_login_required():
    if "admin_user" not in session:
        return False
    return True


def is_admin_record(row_dict):
    email = str(row_dict.get("email", "")).strip().lower()
    return (
        bool(row_dict.get("is_admin"))
        or str(row_dict.get("role", "")).lower() == "admin"
        or (DEFAULT_ADMIN_EMAIL and email == DEFAULT_ADMIN_EMAIL)
    )


def can_check_user_active():
    return "is_active" in get_agent_user_columns()


def build_agent_user_select():
    columns = get_agent_user_columns()
    select_fields = ["id"]
    select_fields.append("name" if "name" in columns else "'' AS name")
    select_fields.append("email" if "email" in columns else "'' AS email")
    select_fields.append("password" if "password" in columns else "'' AS password")
    select_fields.append("role" if "role" in columns else "'user' AS role")
    select_fields.append("is_admin" if "is_admin" in columns else "0 AS is_admin")
    select_fields.append("is_active" if "is_active" in columns else "1 AS is_active")
    return ", ".join(select_fields)


def get_agent_user_by_email(email):
    columns = get_agent_user_columns()
    if not columns:
        return None

    cur = mysql.connection.cursor()
    cur.execute(
        f"SELECT {build_agent_user_select()} FROM agent_users WHERE email=%s LIMIT 1",
        (email,),
    )
    row = cur.fetchone()
    cur.close()
    if not row:
        return None

    return {
        "id": row[0],
        "name": row[1],
        "email": row[2],
        "password": row[3],
        "role": row[4],
        "is_admin": bool(row[5]),
        "is_active": bool(row[6]),
    }


def fetch_all_users():
    columns = get_agent_user_columns()
    if not columns:
        return []

    cur = mysql.connection.cursor()
    cur.execute(
        f"SELECT {build_agent_user_select()} FROM agent_users ORDER BY id DESC"
    )
    rows = cur.fetchall()
    cur.close()
    users = []
    for row in rows:
        users.append(
            {
                "id": row[0],
                "name": row[1] or "",
                "email": row[2] or "",
                "role": row[4] or "user",
                "is_admin": bool(row[5]),
                "is_active": bool(row[6]),
            }
        )
    return users


def get_table_columns(table_name):
    cur = mysql.connection.cursor()
    try:
        cur.execute(f"SHOW COLUMNS FROM {table_name}")
        columns = {row[0] for row in cur.fetchall()}
    except Exception:
        columns = set()
    finally:
        cur.close()
    return columns


def get_agent_call_columns():
    return get_table_columns("agent_calls")


def get_agent_checker_columns():
    return get_table_columns("agent_checker")


def get_agent_user_columns():
    return get_table_columns("agent_users")


def get_wallet_transaction_columns():
    return get_table_columns("agent_wallet_transactions")


def get_support_request_columns():
    return get_table_columns("agent_support_requests")


def get_transfer_settings_columns():
    return get_table_columns("agent_transfer_settings")


def format_dt(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value or "")


def normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def normalize_phone(value):
    raw = str(value or "").strip()
    digits = "".join(ch for ch in raw if ch.isdigit())
    return digits


def parse_schedule_value(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, dt_time(hour=9, minute=0))

    text = str(value).strip()
    if not text:
        return None

    formats = [
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %I:%M %p",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    raise ValueError(f"Invalid schedule time: {text}")


def create_call(phone, schedule_time, user_id, message, contact_name=None):
    columns = get_agent_call_columns()
    insert_columns = ["user_id", "phone", "schedule_time", "status"]
    placeholders = ["%s", "%s", "%s", "%s"]
    values = [user_id, phone, schedule_time, "scheduled"]

    if "message" in columns:
        insert_columns.append("message")
        placeholders.append("%s")
        values.append(message[:255])

    if contact_name and "customer_name" in columns:
        insert_columns.append("customer_name")
        placeholders.append("%s")
        values.append(contact_name[:120])

    cur = mysql.connection.cursor()
    cur.execute(
        f"INSERT INTO agent_calls({', '.join(insert_columns)}) "
        f"VALUES({', '.join(placeholders)})",
        tuple(values),
    )
    call_id = cur.lastrowid
    mysql.connection.commit()
    cur.close()
    return call_id


def update_call_record(call_id, status, message=None):
    columns = get_agent_call_columns()
    updates = ["status=%s"]
    values = [status]

    if message is not None and "message" in columns:
        updates.append("message=%s")
        values.append(message[:255])

    if "updated_at" in columns:
        updates.append("updated_at=NOW()")

    values.append(call_id)

    cur = mysql.connection.cursor()
    cur.execute(
        f"UPDATE agent_calls SET {', '.join(updates)} WHERE id=%s",
        tuple(values),
    )
    mysql.connection.commit()
    cur.close()


def fetch_calls_for_user(user_id, limit=None):
    columns = get_agent_call_columns()
    select_fields = ["id", "phone", "schedule_time", "status"]

    if "message" in columns:
        select_fields.append("message")
    else:
        select_fields.append("'' AS message")

    if "updated_at" in columns:
        select_fields.append("updated_at")
    else:
        select_fields.append("schedule_time AS updated_at")

    if "customer_name" in columns:
        select_fields.append("customer_name")
    else:
        select_fields.append("'' AS customer_name")

    query = (
        f"SELECT {', '.join(select_fields)} "
        "FROM agent_calls WHERE user_id=%s ORDER BY schedule_time DESC, id DESC"
    )

    if limit:
        query += f" LIMIT {int(limit)}"

    cur = mysql.connection.cursor()
    cur.execute(query, (user_id,))
    rows = cur.fetchall()
    cur.close()

    calls = []
    for row in rows:
        calls.append(
            {
                "id": row[0],
                "phone": row[1],
                "time": format_dt(row[2]),
                "status": row[3],
                "message": row[4] or "",
                "updated_at": format_dt(row[5]),
                "customer_name": row[6] or "",
            }
        )

    return calls


def get_current_agent_profile(user_id):
    columns = get_agent_checker_columns()
    if not columns:
        return None

    cur = mysql.connection.cursor()
    cur.execute(
        """
        SELECT id, credential_id, assigned_number, sip_user, sip_pass, active, is_approved, approved_at
        FROM agent_checker
        WHERE approved_user_id=%s AND is_approved=1
        ORDER BY approved_at DESC, id DESC
        LIMIT 1
        """,
        (user_id,),
    )
    row = cur.fetchone()
    cur.close()

    if not row:
        return None

    return {
        "id": row[0],
        "credential_id": row[1],
        "assigned_number": row[2],
        "sip_user": row[3],
        "sip_pass": row[4],
        "active": bool(row[5]),
        "is_approved": bool(row[6]),
        "approved_at": format_dt(row[7]),
    }


def get_transfer_settings(user_id):
    columns = get_transfer_settings_columns()
    if not columns:
        return {
            "enabled": False,
            "transfer_number": "",
            "table_ready": False,
        }

    cur = mysql.connection.cursor()
    cur.execute(
        """
        SELECT transfer_number, is_enabled
        FROM agent_transfer_settings
        WHERE user_id=%s
        LIMIT 1
        """,
        (user_id,),
    )
    row = cur.fetchone()
    cur.close()

    if not row:
        return {
            "enabled": False,
            "transfer_number": "",
            "table_ready": True,
        }

    return {
        "enabled": bool(row[1]),
        "transfer_number": row[0] or "",
        "table_ready": True,
    }


def save_transfer_settings(user_id, transfer_number, is_enabled):
    columns = get_transfer_settings_columns()
    if not columns:
        return False, "Please create `agent_transfer_settings` table first."

    cur = mysql.connection.cursor()
    cur.execute("SELECT id FROM agent_transfer_settings WHERE user_id=%s LIMIT 1", (user_id,))
    row = cur.fetchone()

    if row:
        cur.execute(
            """
            UPDATE agent_transfer_settings
            SET transfer_number=%s, is_enabled=%s, updated_at=NOW()
            WHERE user_id=%s
            """,
            (transfer_number, 1 if is_enabled else 0, user_id),
        )
    else:
        cur.execute(
            """
            INSERT INTO agent_transfer_settings(user_id, transfer_number, is_enabled)
            VALUES(%s, %s, %s)
            """,
            (user_id, transfer_number, 1 if is_enabled else 0),
        )

    mysql.connection.commit()
    cur.close()
    return True, "Transfer settings saved successfully."


def approve_agent_for_user(user_id, credential_id, credential_password):
    if not get_agent_checker_columns():
        return False, "Table `agent_checker` not found. Please create it first."

    cur = mysql.connection.cursor()
    cur.execute(
        """
        SELECT id, credential_id, credential_password, assigned_number, sip_user, sip_pass, active, approved_user_id
        FROM agent_checker
        WHERE credential_id=%s
        LIMIT 1
        """,
        (credential_id,),
    )
    row = cur.fetchone()

    if not row:
        cur.close()
        return False, "Credential ID not found in agent_checker."

    if str(row[2]) != str(credential_password):
        cur.close()
        return False, "Credential password does not match."

    if not row[6]:
        cur.close()
        return False, "This agent credential is inactive."

    if row[7] and int(row[7]) != int(user_id):
        cur.close()
        return False, "This credential is already approved for another user."

    cur.execute(
        """
        UPDATE agent_checker
        SET is_approved=0, approved_user_id=NULL, approved_at=NULL
        WHERE approved_user_id=%s
        """,
        (user_id,),
    )
    cur.execute(
        """
        UPDATE agent_checker
        SET is_approved=1, approved_user_id=%s, approved_at=NOW()
        WHERE id=%s
        """,
        (user_id, row[0]),
    )
    mysql.connection.commit()
    cur.close()
    return True, "Agent approved successfully."


def fetch_live_call_records():
    if not CALL_RECORD_AUTH_ID or not CALL_RECORD_AUTH_TOKEN or not CALL_RECORD_API_URL:
        raise ValueError("Call record API credentials are missing.")

    req = Request(
        CALL_RECORD_API_URL,
        headers={
            "X-Auth-ID": CALL_RECORD_AUTH_ID,
            "X-Auth-Token": CALL_RECORD_AUTH_TOKEN,
            "Content-Type": "application/json",
        },
    )

    try:
        with urlopen(req, timeout=20) as response:
            payload = response.read().decode("utf-8")
    except url_error.URLError as exc:
        raise ValueError(f"Unable to load call detail records: {exc}") from exc

    import json

    data = json.loads(payload)
    return data.get("data", []) if isinstance(data, dict) else []


def get_wallet_summary(user_id, used_amount):
    columns = get_wallet_transaction_columns()
    if not columns:
        return {
            "wallet_enabled": False,
            "total_recharged": 0.0,
            "pending_recharge": 0.0,
            "remaining_balance": 0.0,
            "used_amount": round(used_amount, 2),
        }

    cur = mysql.connection.cursor()
    cur.execute(
        """
        SELECT COALESCE(SUM(amount), 0)
        FROM agent_wallet_transactions
        WHERE user_id=%s AND transaction_type='recharge' AND payment_status='paid'
        """,
        (user_id,),
    )
    total_recharged = float(cur.fetchone()[0] or 0)

    cur.execute(
        """
        SELECT COALESCE(SUM(amount), 0)
        FROM agent_wallet_transactions
        WHERE user_id=%s AND transaction_type='recharge' AND payment_status='pending'
        """,
        (user_id,),
    )
    pending_recharge = float(cur.fetchone()[0] or 0)
    cur.close()

    return {
        "wallet_enabled": True,
        "total_recharged": round(total_recharged, 2),
        "pending_recharge": round(pending_recharge, 2),
        "remaining_balance": round(total_recharged - used_amount, 2),
        "used_amount": round(used_amount, 2),
    }


def get_wallet_transactions_for_user(user_id, limit=12):
    columns = get_wallet_transaction_columns()
    if not columns:
        return []

    select_fields = ["id"]
    select_fields.append("payment_reference" if "payment_reference" in columns else "'' AS payment_reference")
    select_fields.append("amount" if "amount" in columns else "0 AS amount")
    select_fields.append("transaction_type" if "transaction_type" in columns else "'recharge' AS transaction_type")
    select_fields.append("payment_status" if "payment_status" in columns else "'pending' AS payment_status")
    select_fields.append("gateway_payment_id" if "gateway_payment_id" in columns else "'' AS gateway_payment_id")
    select_fields.append("note" if "note" in columns else "'' AS note")
    select_fields.append("created_at" if "created_at" in columns else "NOW() AS created_at")
    select_fields.append("paid_at" if "paid_at" in columns else "NULL AS paid_at")

    cur = mysql.connection.cursor()
    cur.execute(
        f"""
        SELECT {', '.join(select_fields)}
        FROM agent_wallet_transactions
        WHERE user_id=%s
        ORDER BY id DESC
        LIMIT %s
        """,
        (user_id, limit),
    )
    rows = cur.fetchall()
    cur.close()

    transactions = []
    for row in rows:
        transactions.append(
            {
                "id": row[0],
                "payment_reference": row[1] or "",
                "amount": float(row[2] or 0),
                "transaction_type": row[3] or "",
                "payment_status": row[4] or "pending",
                "gateway_payment_id": row[5] or "",
                "note": row[6] or "",
                "created_at": format_dt(row[7]),
                "paid_at": format_dt(row[8]) if row[8] else "",
            }
        )

    return transactions


def set_ui_message(message, message_type="info"):
    session["ui_message"] = message
    session["ui_message_type"] = message_type


def pop_ui_message():
    message = session.pop("ui_message", "")
    message_type = session.pop("ui_message_type", "")
    return message, message_type


def create_wallet_recharge_request(user_id, amount, note=None):
    columns = get_wallet_transaction_columns()
    if not columns:
        return False, "Please create `agent_wallet_transactions` table first.", None

    payment_reference = f"AWT-{uuid.uuid4().hex[:16].upper()}"
    note = note or f"Wallet recharge request of Rs {amount:.2f}"

    insert_columns = []
    values = []

    field_map = [
        ("user_id", user_id),
        ("payment_reference", payment_reference),
        ("amount", amount),
        ("transaction_type", "recharge"),
        ("payment_status", "pending"),
        ("note", note),
        ("gateway_name", "razorpay"),
    ]

    for column_name, value in field_map:
        if column_name in columns:
            insert_columns.append(column_name)
            values.append(value)

    if "user_id" not in insert_columns or "amount" not in insert_columns:
        return False, "Wallet table structure is incomplete. Please update `agent_wallet_transactions` table.", None

    cur = mysql.connection.cursor()
    placeholders = ", ".join(["%s"] * len(insert_columns))
    cur.execute(
        f"INSERT INTO agent_wallet_transactions({', '.join(insert_columns)}) VALUES({placeholders})",
        tuple(values),
    )
    mysql.connection.commit()
    cur.close()
    return True, "Recharge request created.", payment_reference


def update_wallet_gateway_link(payment_reference, gateway_link_id, gateway_link_url, gateway_payload=""):
    columns = get_wallet_transaction_columns()
    if not columns:
        return

    updates = []
    values = []
    if "gateway_link_id" in columns:
        updates.append("gateway_link_id=%s")
        values.append(gateway_link_id)
    if "gateway_link_url" in columns:
        updates.append("gateway_link_url=%s")
        values.append(gateway_link_url)
    if "gateway_payload" in columns:
        updates.append("gateway_payload=%s")
        values.append(gateway_payload[:5000])
    if "updated_at" in columns:
        updates.append("updated_at=NOW()")

    if not updates:
        return

    cur = mysql.connection.cursor()
    cur.execute(
        f"UPDATE agent_wallet_transactions SET {', '.join(updates)} WHERE payment_reference=%s",
        tuple(values + [payment_reference]),
    )
    mysql.connection.commit()
    cur.close()


def get_wallet_transaction(payment_reference):
    columns = get_wallet_transaction_columns()
    if not columns:
        return None

    select_fields = ["id"]
    select_fields.append("user_id" if "user_id" in columns else "0 AS user_id")
    select_fields.append("payment_reference" if "payment_reference" in columns else "'' AS payment_reference")
    select_fields.append("amount" if "amount" in columns else "0 AS amount")
    select_fields.append("payment_status" if "payment_status" in columns else "'pending' AS payment_status")
    select_fields.append("gateway_payment_id" if "gateway_payment_id" in columns else "'' AS gateway_payment_id")
    select_fields.append("gateway_link_id" if "gateway_link_id" in columns else "'' AS gateway_link_id")
    select_fields.append("gateway_link_url" if "gateway_link_url" in columns else "'' AS gateway_link_url")

    cur = mysql.connection.cursor()
    cur.execute(
        f"""
        SELECT {', '.join(select_fields)}
        FROM agent_wallet_transactions
        WHERE payment_reference=%s
        LIMIT 1
        """,
        (payment_reference,),
    )
    row = cur.fetchone()
    cur.close()
    if not row:
        return None

    return {
        "id": row[0],
        "user_id": row[1],
        "payment_reference": row[2],
        "amount": float(row[3] or 0),
        "payment_status": row[4] or "pending",
        "gateway_payment_id": row[5] or "",
        "gateway_link_id": row[6] or "",
        "gateway_link_url": row[7] or "",
    }


def razorpay_enabled():
    return bool(RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET)


def razorpay_request(method, path, payload=None):
    if not razorpay_enabled():
        raise ValueError("Razorpay keys are missing. Add RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET in .env")

    url = f"{RAZORPAY_API_URL.rstrip('/')}/{path.lstrip('/')}"
    auth_token = base64.b64encode(f"{RAZORPAY_KEY_ID}:{RAZORPAY_KEY_SECRET}".encode("utf-8")).decode("utf-8")
    request_data = None
    headers = {
        "Authorization": f"Basic {auth_token}",
        "Content-Type": "application/json",
    }

    if payload is not None:
        request_data = json.dumps(payload).encode("utf-8")

    req = Request(url, data=request_data, headers=headers, method=method.upper())
    with urlopen(req, timeout=20) as response:
        return json.loads(response.read().decode("utf-8"))


def create_razorpay_payment_link(user_id, payment_reference, amount):
    callback_url = url_for("wallet_payment_callback", _external=True)
    payload = {
        "amount": int(round(amount * 100)),
        "currency": "INR",
        "accept_partial": False,
        "description": f"Wallet recharge for user {user_id}",
        "reference_id": payment_reference,
        "callback_url": callback_url,
        "callback_method": "get",
        "notes": {
            "user_id": str(user_id),
            "payment_reference": payment_reference,
        },
    }
    response = razorpay_request("POST", "/payment_links", payload)
    return {
        "id": response.get("id", ""),
        "short_url": response.get("short_url", ""),
        "status": response.get("status", ""),
        "payload": json.dumps(response),
    }


def verify_razorpay_callback_signature(payment_link_id, payment_reference, payment_status, payment_id, signature):
    body = f"{payment_link_id}|{payment_reference}|{payment_status}|{payment_id}"
    expected = hmac.new(
        RAZORPAY_KEY_SECRET.encode("utf-8"),
        body.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return hmac.compare_digest(expected, signature or "")


def fetch_razorpay_payment_link(payment_link_id):
    return razorpay_request("GET", f"/payment_links/{payment_link_id}")


def update_wallet_transaction_status(
    payment_reference,
    payment_status,
    gateway_payment_id="",
    gateway_order_id="",
    gateway_signature="",
    gateway_payload="",
):
    columns = get_wallet_transaction_columns()
    if not columns:
        return False, "Please create `agent_wallet_transactions` table first."

    if payment_status not in {"pending", "paid", "failed"}:
        return False, "Invalid payment status."

    cur = mysql.connection.cursor()
    cur.execute(
        """
        SELECT id, payment_status
        FROM agent_wallet_transactions
        WHERE payment_reference=%s
        LIMIT 1
        """,
        (payment_reference,),
    )
    row = cur.fetchone()

    if not row:
        cur.close()
        return False, "Payment reference not found."

    if row[1] == "paid" and payment_status == "failed":
        cur.close()
        return False, "Paid transaction cannot be marked failed."

    cur.execute(
        """
        UPDATE agent_wallet_transactions
        SET payment_status=%s,
            gateway_payment_id=%s,
            gateway_order_id=%s,
            gateway_signature=%s,
            gateway_payload=%s,
            paid_at=CASE WHEN %s='paid' THEN NOW() ELSE paid_at END,
            updated_at=NOW()
        WHERE payment_reference=%s
        """,
        (
            payment_status,
            gateway_payment_id,
            gateway_order_id,
            gateway_signature,
            gateway_payload[:5000],
            payment_status,
            payment_reference,
        ),
    )
    mysql.connection.commit()
    cur.close()
    return True, "Transaction updated successfully."


def build_wallet_redirect_url(payment_reference, amount):
    success_url = url_for("wallet_payment_success", _external=True)
    failed_url = url_for("wallet_payment_failed", _external=True)
    query = urlencode(
        {
            "amount": f"{amount:.2f}",
            "payment_reference": payment_reference,
            "redirect_success": success_url,
            "redirect_failed": failed_url,
        }
    )
    separator = "&" if "?" in WALLET_RECHARGE_URL else "?"
    return f"{WALLET_RECHARGE_URL}{separator}{query}"


def get_filtered_call_records(user_id):
    scheduled_calls = fetch_calls_for_user(user_id)
    approved_agent = get_current_agent_profile(user_id)
    scheduled_numbers = {normalize_phone(call["phone"]) for call in scheduled_calls if call.get("phone")}
    call_names = {
        normalize_phone(call["phone"]): (call.get("customer_name") or "")
        for call in scheduled_calls
        if call.get("phone")
    }

    if not scheduled_numbers:
        wallet = get_wallet_summary(user_id, 0)
        return {
            "summary": {
                "total_records": 0,
                "talk_minutes": 0,
                "billable_minutes": 0,
                "billed_amount": 0,
                "wallet_recharged": wallet["total_recharged"],
                "wallet_pending": wallet["pending_recharge"],
                "wallet_remaining": wallet["remaining_balance"],
            },
            "customer_summary": [],
            "records": [],
            "wallet": wallet,
            "transactions": get_wallet_transactions_for_user(user_id),
        }

    assigned_number = normalize_phone(approved_agent["assigned_number"]) if approved_agent else ""
    records = []
    customer_summary = defaultdict(
        lambda: {
            "customer_name": "",
            "phone": "",
            "call_count": 0,
            "talk_minutes": 0.0,
            "billable_minutes": 0,
            "billed_amount": 0.0,
        }
    )

    for record in fetch_live_call_records():
        destination = normalize_phone(record.get("destination_number"))
        caller = normalize_phone(record.get("caller_id_number"))

        if destination not in scheduled_numbers:
            continue

        if assigned_number and caller and caller != assigned_number:
            continue

        duration_seconds = int(record.get("duration") or 0)
        talk_minutes = round(duration_seconds / 60, 2)
        billable_minutes = math.ceil(duration_seconds / 60) if duration_seconds > 0 else 0
        billed_amount = round(billable_minutes * SELL_RATE_PER_MINUTE, 2)
        customer_name = call_names.get(destination, "")
        status_text = record.get("hangup_cause") or "UNKNOWN"

        item = {
            "id": record.get("id"),
            "caller": record.get("caller_id_number") or "",
            "phone": record.get("destination_number") or "",
            "customer_name": customer_name,
            "duration_seconds": duration_seconds,
            "talk_minutes": talk_minutes,
            "billable_minutes": billable_minutes,
            "billed_amount": billed_amount,
            "status": status_text,
            "start_time": record.get("start_time") or "",
        }
        records.append(item)

        summary_key = destination
        customer_summary[summary_key]["customer_name"] = customer_name or customer_summary[summary_key]["customer_name"]
        customer_summary[summary_key]["phone"] = item["phone"]
        customer_summary[summary_key]["call_count"] += 1
        customer_summary[summary_key]["talk_minutes"] = round(
            customer_summary[summary_key]["talk_minutes"] + talk_minutes, 2
        )
        customer_summary[summary_key]["billable_minutes"] += billable_minutes
        customer_summary[summary_key]["billed_amount"] = round(
            customer_summary[summary_key]["billed_amount"] + billed_amount, 2
        )

    records.sort(key=lambda row: row["start_time"], reverse=True)
    summary_rows = sorted(customer_summary.values(), key=lambda row: row["billed_amount"], reverse=True)
    total_talk_minutes = round(sum(row["talk_minutes"] for row in records), 2)
    total_billable_minutes = sum(row["billable_minutes"] for row in records)
    total_billed_amount = round(sum(row["billed_amount"] for row in records), 2)
    wallet = get_wallet_summary(user_id, total_billed_amount)

    return {
        "summary": {
            "total_records": len(records),
            "talk_minutes": total_talk_minutes,
            "billable_minutes": total_billable_minutes,
            "billed_amount": total_billed_amount,
            "wallet_recharged": wallet["total_recharged"],
            "wallet_pending": wallet["pending_recharge"],
            "wallet_remaining": wallet["remaining_balance"],
        },
        "customer_summary": summary_rows,
        "records": records,
        "wallet": wallet,
        "transactions": get_wallet_transactions_for_user(user_id),
    }


def get_dashboard_payload(user_id):
    calls = fetch_calls_for_user(user_id, limit=10)
    all_calls = fetch_calls_for_user(user_id)
    approved_agent = get_current_agent_profile(user_id)
    transfer_settings = get_transfer_settings(user_id)
    try:
        call_record_data = get_filtered_call_records(user_id)
    except ValueError:
        call_record_data = {
            "summary": {
                "talk_minutes": 0,
                "billable_minutes": 0,
                "billed_amount": 0,
                "wallet_recharged": 0,
                "wallet_remaining": 0,
            }
        }

    live_statuses = {"calling", "ringing"}
    no_response_statuses = {"not_picked", "no_response"}
    failed_statuses = {"failed", "error", "rejected", "invalid"}

    stats = {
        "total": len(all_calls),
        "scheduled": sum(1 for call in all_calls if call["status"] == "scheduled"),
        "live": sum(1 for call in all_calls if call["status"] in live_statuses),
        "answered": sum(1 for call in all_calls if call["status"] == "answered"),
        "no_response": sum(1 for call in all_calls if call["status"] in no_response_statuses),
        "failed": sum(1 for call in all_calls if call["status"] in failed_statuses),
        "approved": bool(approved_agent),
        "assigned_number": approved_agent["assigned_number"] if approved_agent else "",
        "credential_id": approved_agent["credential_id"] if approved_agent else "",
        "talk_minutes": call_record_data["summary"]["talk_minutes"],
        "wallet_remaining": call_record_data["summary"]["wallet_remaining"],
        "transfer_enabled": transfer_settings["enabled"],
        "transfer_number": transfer_settings["transfer_number"],
    }

    if stats["total"]:
        stats["success_rate"] = round((stats["answered"] / stats["total"]) * 100, 1)
    else:
        stats["success_rate"] = 0

    return {
        "stats": stats,
        "calls": calls,
        "agent": approved_agent,
        "transfer_settings": transfer_settings,
        "call_record_summary": call_record_data["summary"],
    }


def fetch_all_credentials():
    columns = get_agent_checker_columns()
    if not columns:
        return []

    cur = mysql.connection.cursor()
    cur.execute(
        """
        SELECT id, credential_id, assigned_number, sip_user, active, is_approved, approved_user_id, approved_at
        FROM agent_checker
        ORDER BY id DESC
        """
    )
    rows = cur.fetchall()
    cur.close()
    credentials = []
    for row in rows:
        credentials.append(
            {
                "id": row[0],
                "credential_id": row[1] or "",
                "assigned_number": row[2] or "",
                "sip_user": row[3] or "",
                "active": bool(row[4]),
                "is_approved": bool(row[5]),
                "approved_user_id": row[6] or "",
                "approved_at": format_dt(row[7]) if row[7] else "",
            }
        )
    return credentials


def fetch_all_wallet_transactions(limit=200):
    columns = get_wallet_transaction_columns()
    if not columns:
        return []

    cur = mysql.connection.cursor()
    cur.execute(
        """
        SELECT wt.id, wt.user_id, u.name, u.email, wt.payment_reference, wt.amount, wt.transaction_type,
               wt.payment_status, wt.gateway_name, wt.gateway_link_id, wt.gateway_payment_id, wt.created_at, wt.paid_at
        FROM agent_wallet_transactions wt
        LEFT JOIN agent_users u ON u.id = wt.user_id
        ORDER BY wt.id DESC
        LIMIT %s
        """,
        (limit,),
    )
    rows = cur.fetchall()
    cur.close()

    items = []
    for row in rows:
        items.append(
            {
                "id": row[0],
                "user_id": row[1],
                "user_name": row[2] or "",
                "user_email": row[3] or "",
                "payment_reference": row[4] or "",
                "amount": float(row[5] or 0),
                "transaction_type": row[6] or "",
                "payment_status": row[7] or "pending",
                "gateway_name": row[8] or "",
                "gateway_link_id": row[9] or "",
                "gateway_payment_id": row[10] or "",
                "created_at": format_dt(row[11]),
                "paid_at": format_dt(row[12]) if row[12] else "",
            }
        )
    return items


def create_agent_credential(payload):
    columns = get_agent_checker_columns()
    if not columns:
        return False, "agent_checker table not found."

    insert_columns = []
    values = []

    field_map = [
        ("credential_id", payload.get("credential_id", "").strip()),
        ("credential_password", payload.get("credential_password", "").strip()),
        ("assigned_number", payload.get("assigned_number", "").strip()),
        ("sip_user", payload.get("sip_user", "").strip()),
        ("sip_pass", payload.get("sip_pass", "").strip()),
    ]

    for column_name, value in field_map:
        if column_name in columns:
            insert_columns.append(column_name)
            values.append(value)

    if "active" in columns:
        insert_columns.append("active")
        values.append(1 if payload.get("active", True) else 0)

    if "is_approved" in columns:
        insert_columns.append("is_approved")
        values.append(0)

    if not insert_columns or "credential_id" not in insert_columns:
        return False, "Required credential columns are missing in agent_checker."

    placeholders = ", ".join(["%s"] * len(insert_columns))
    cur = mysql.connection.cursor()
    cur.execute(
        f"INSERT INTO agent_checker({', '.join(insert_columns)}) VALUES({placeholders})",
        tuple(values),
    )
    mysql.connection.commit()
    cur.close()
    return True, "Credential added successfully."


def create_support_request(user_id, page_name, issue, expected_outcome, note):
    columns = get_support_request_columns()
    if not columns:
        return False, "Please create `agent_support_requests` table first."

    cur = mysql.connection.cursor()
    cur.execute(
        """
        INSERT INTO agent_support_requests(user_id, page_name, issue, expected_outcome, note, status)
        VALUES(%s, %s, %s, %s, %s, 'open')
        """,
        (user_id, page_name[:50], issue[:255], expected_outcome[:255], note[:1000]),
    )
    mysql.connection.commit()
    cur.close()
    return True, "Support request created."


def fetch_support_requests(limit=50):
    columns = get_support_request_columns()
    if not columns:
        return []

    cur = mysql.connection.cursor()
    cur.execute(
        """
        SELECT sr.id, sr.user_id, u.name, u.email, sr.page_name, sr.issue, sr.expected_outcome, sr.note, sr.status, sr.created_at
        FROM agent_support_requests sr
        LEFT JOIN agent_users u ON u.id = sr.user_id
        ORDER BY sr.id DESC
        LIMIT %s
        """,
        (limit,),
    )
    rows = cur.fetchall()
    cur.close()
    requests_list = []
    for row in rows:
        requests_list.append(
            {
                "id": row[0],
                "user_id": row[1],
                "user_name": row[2] or "",
                "user_email": row[3] or "",
                "page_name": row[4] or "",
                "issue": row[5] or "",
                "expected_outcome": row[6] or "",
                "note": row[7] or "",
                "status": row[8] or "open",
                "created_at": format_dt(row[9]),
            }
        )
    return requests_list


def get_admin_dashboard_payload():
    return {
        "users": fetch_all_users(),
        "credentials": fetch_all_credentials(),
        "support_requests": fetch_support_requests(),
    }


def get_admin_payments_payload():
    return {
        "transactions": fetch_all_wallet_transactions(),
    }


def schedule_call_job(call_id, phone, schedule_time):
    scheduler.add_job(
        run_call,
        "date",
        run_date=schedule_time,
        args=[call_id, phone],
        id=f"call-{call_id}",
        replace_existing=True,
    )


def restore_scheduled_calls():
    columns = get_agent_call_columns()
    if not columns or "schedule_time" not in columns or "status" not in columns:
        return

    cur = mysql.connection.cursor()
    try:
        cur.execute(
            """
            SELECT id, phone, schedule_time
            FROM agent_calls
            WHERE status=%s
            ORDER BY schedule_time ASC, id ASC
            """,
            ("scheduled",),
        )
        rows = cur.fetchall()
    finally:
        cur.close()

    now = datetime.now()
    for row in rows:
        call_id, phone, schedule_time = row
        if not phone or schedule_time is None:
            continue

        if schedule_time <= now:
            update_call_record(call_id, "rejected", "Scheduled time passed before the call could be started.")
            continue

        schedule_call_job(call_id, phone, schedule_time)


def start_scheduler_once():
    global _scheduler_bootstrapped

    if _scheduler_bootstrapped or not SCHEDULER_ENABLED:
        return

    with app.app_context():
        if not scheduler.running:
            scheduler.start()

        restore_scheduled_calls()
        _scheduler_bootstrapped = True


def run_call(call_id, phone):
    with app.app_context():
        update_call_record(call_id, "calling", "Dispatching AI agent.")

    command = [sys.executable, "make_call.py", "--to", phone, "--call-id", str(call_id)]

    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        cwd=os.path.dirname(os.path.abspath(__file__)),
    )

    if result.returncode != 0:
        error_text = (result.stderr or result.stdout or "Call dispatch failed.").strip()
        with app.app_context():
            update_call_record(call_id, "rejected", error_text)


def require_approved_agent():
    if not login_required():
        return None, (jsonify({"ok": False, "error": "Unauthorized"}), 401)

    agent = get_current_agent_profile(session["user"])
    if not agent:
        return None, (
            jsonify({"ok": False, "error": "Please approve your agent credential from Settings first."}),
            403,
        )

    return agent, None


def schedule_single_call(user_id, phone, schedule_time, contact_name=None):
    call_id = create_call(
        phone=phone,
        schedule_time=schedule_time,
        user_id=user_id,
        message="Waiting for scheduled time.",
        contact_name=contact_name,
    )
    schedule_call_job(call_id, phone, schedule_time)
    return call_id


def parse_excel_rows(uploaded_file, default_schedule_time):
    if load_workbook is None:
        raise ValueError("openpyxl is not installed. Please install requirements first.")

    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalize_header(value) for value in rows[0]]
    phone_index = next((i for i, value in enumerate(headers) if value in PHONE_HEADERS), None)
    time_index = next((i for i, value in enumerate(headers) if value in TIME_HEADERS), None)
    name_index = next((i for i, value in enumerate(headers) if value in NAME_HEADERS), None)

    if phone_index is None:
        raise ValueError("Excel file must contain a phone column like phone, mobile, or phone_number.")

    parsed_rows = []

    for raw_row in rows[1:]:
        if not raw_row:
            continue

        phone = str(raw_row[phone_index] or "").strip()
        if not phone:
            continue

        row_time = raw_row[time_index] if time_index is not None and time_index < len(raw_row) else None
        schedule_time = parse_schedule_value(row_time) if row_time not in (None, "") else default_schedule_time
        if schedule_time is None:
            raise ValueError(f"Schedule time missing for phone {phone}.")

        contact_name = ""
        if name_index is not None and name_index < len(raw_row):
            contact_name = str(raw_row[name_index] or "").strip()

        parsed_rows.append(
            {
                "phone": phone,
                "schedule_time": schedule_time,
                "customer_name": contact_name,
            }
        )

    return parsed_rows


@app.route("/")
def home():
    if "user" in session:
        return redirect("/dashboard")
    return redirect("/login")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        user = get_agent_user_by_email(email)
        if user and user["password"] == password and user["is_active"]:
            session["user"] = user["id"]
            return redirect("/dashboard")
        set_ui_message("Invalid email or password. Please try again.", "error")

    message, message_type = pop_ui_message()
    return render_template("login.html", message=message, message_type=message_type)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]
        user = get_agent_user_by_email(email)

        if user and user["password"] == password and user["is_active"] and is_admin_record(user):
            session["admin_user"] = user["id"]
            session["admin_name"] = user["name"]
            return redirect("/admin/dashboard")

        set_ui_message("Admin login failed. Check role, active status, email, and password.", "error")

    message, message_type = pop_ui_message()
    return render_template("admin_login.html", message=message, message_type=message_type)


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_user", None)
    session.pop("admin_name", None)
    return redirect("/admin/login")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]

        columns = get_agent_user_columns()
        cur = mysql.connection.cursor()
        cur.execute("SELECT id FROM agent_users WHERE email=%s LIMIT 1", (email,))
        existing_user = cur.fetchone()
        if existing_user:
            cur.close()
            set_ui_message("This email is already registered. Please login instead.", "error")
            return redirect("/register")

        insert_columns = ["name", "email", "password"]
        placeholders = ["%s", "%s", "%s"]
        values = [name, email, password]
        if "role" in columns:
            insert_columns.append("role")
            placeholders.append("%s")
            values.append("user")
        if "is_admin" in columns:
            insert_columns.append("is_admin")
            placeholders.append("%s")
            values.append(0)
        if "is_active" in columns:
            insert_columns.append("is_active")
            placeholders.append("%s")
            values.append(1)

        cur.execute(
            f"INSERT INTO agent_users({', '.join(insert_columns)}) VALUES({', '.join(placeholders)})",
            tuple(values),
        )
        mysql.connection.commit()
        cur.close()

        set_ui_message("Registration completed. Please login to continue.", "success")
        return redirect("/login")

    message, message_type = pop_ui_message()
    return render_template("register.html", message=message, message_type=message_type)


@app.route("/support_request", methods=["POST"])
def support_request():
    if not login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    data = request.get_json(silent=True) or {}
    issue = (data.get("issue") or "").strip()
    expected_outcome = (data.get("expected_outcome") or "").strip()
    page_name = (data.get("page_name") or "dashboard").strip()
    note = (data.get("note") or "").strip()

    if not issue or not expected_outcome:
        return jsonify({"ok": False, "error": "Issue and expected outcome are required."}), 400

    success, message = create_support_request(session["user"], page_name, issue, expected_outcome, note)
    return jsonify({"ok": success, "message": message}), 200 if success else 400


@app.route("/dashboard")
def dashboard():
    if not login_required():
        return redirect("/login")

    payload = get_dashboard_payload(session["user"])
    message, message_type = pop_ui_message()
    return render_template("dashboard.html", payload=payload, message=message, message_type=message_type)


@app.route("/admin/dashboard")
def admin_dashboard():
    if not admin_login_required():
        return redirect("/admin/login")

    payload = get_admin_dashboard_payload()
    message, message_type = pop_ui_message()
    return render_template("admin_dashboard.html", payload=payload, message=message, message_type=message_type)


@app.route("/admin/payments")
def admin_payments():
    if not admin_login_required():
        return redirect("/admin/login")

    payload = get_admin_payments_payload()
    message, message_type = pop_ui_message()
    return render_template("admin_payments.html", payload=payload, message=message, message_type=message_type)


@app.route("/admin/dashboard_data")
def admin_dashboard_data():
    if not admin_login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    return jsonify({"ok": True, "payload": get_admin_dashboard_payload()})


@app.route("/admin/payments_data")
def admin_payments_data():
    if not admin_login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    return jsonify({"ok": True, "payload": get_admin_payments_payload()})


@app.route("/admin/user/<int:user_id>/update", methods=["POST"])
def admin_update_user(user_id):
    if not admin_login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    columns = get_agent_user_columns()
    if not columns:
        return jsonify({"ok": False, "error": "agent_users table not found."}), 400

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    role = (data.get("role") or "user").strip().lower()
    is_active = 1 if data.get("is_active", True) else 0

    updates = []
    values = []
    if "name" in columns and name:
        updates.append("name=%s")
        values.append(name)
    if "email" in columns and email:
        updates.append("email=%s")
        values.append(email)
    if "role" in columns:
        updates.append("role=%s")
        values.append("admin" if role == "admin" else "user")
    if "is_admin" in columns:
        updates.append("is_admin=%s")
        values.append(1 if role == "admin" else 0)
    if "is_active" in columns:
        updates.append("is_active=%s")
        values.append(is_active)

    if not updates:
        return jsonify({"ok": False, "error": "No editable admin columns found on agent_users."}), 400

    values.append(user_id)
    cur = mysql.connection.cursor()
    cur.execute(f"UPDATE agent_users SET {', '.join(updates)} WHERE id=%s", tuple(values))
    mysql.connection.commit()
    cur.close()
    return jsonify({"ok": True, "message": "User updated successfully."})


@app.route("/admin/credential/<int:credential_row_id>/update", methods=["POST"])
def admin_update_credential(credential_row_id):
    if not admin_login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    if not get_agent_checker_columns():
        return jsonify({"ok": False, "error": "agent_checker table not found."}), 400

    data = request.get_json(silent=True) or {}
    active = 1 if data.get("active", True) else 0
    approved_user_id = data.get("approved_user_id")
    action = (data.get("action") or "").strip().lower()

    cur = mysql.connection.cursor()
    cur.execute("UPDATE agent_checker SET active=%s WHERE id=%s", (active, credential_row_id))

    if action == "revoke":
        cur.execute(
            """
            UPDATE agent_checker
            SET is_approved=0, approved_user_id=NULL, approved_at=NULL
            WHERE id=%s
            """,
            (credential_row_id,),
        )
    elif approved_user_id:
        cur.execute(
            """
            UPDATE agent_checker
            SET is_approved=1, approved_user_id=%s, approved_at=NOW()
            WHERE id=%s
            """,
            (approved_user_id, credential_row_id),
        )

    mysql.connection.commit()
    cur.close()
    return jsonify({"ok": True, "message": "Credential updated successfully."})


@app.route("/admin/credential/create", methods=["POST"])
def admin_create_credential():
    if not admin_login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    data = request.get_json(silent=True) or {}
    if not (data.get("credential_id") or "").strip():
        return jsonify({"ok": False, "error": "Credential ID is required."}), 400

    success, message = create_agent_credential(data)
    return jsonify({"ok": success, "message": message}), 200 if success else 400


@app.route("/admin/support/<int:ticket_id>/update", methods=["POST"])
def admin_update_support(ticket_id):
    if not admin_login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    if not get_support_request_columns():
        return jsonify({"ok": False, "error": "agent_support_requests table not found."}), 400

    data = request.get_json(silent=True) or {}
    status = (data.get("status") or "open").strip().lower()
    if status not in {"open", "in_progress", "resolved", "closed"}:
        return jsonify({"ok": False, "error": "Invalid support status."}), 400

    cur = mysql.connection.cursor()
    cur.execute(
        "UPDATE agent_support_requests SET status=%s, updated_at=NOW() WHERE id=%s",
        (status, ticket_id),
    )
    mysql.connection.commit()
    cur.close()
    return jsonify({"ok": True, "message": "Support ticket updated successfully."})


@app.route("/dashboard_data")
def dashboard_data():
    if not login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    return jsonify(get_dashboard_payload(session["user"]))


@app.route("/call-detail-record")
def call_detail_record():
    if not login_required():
        return redirect("/login")

    try:
        payload = get_filtered_call_records(session["user"])
        load_error = ""
    except ValueError as exc:
        payload = {
            "summary": {
                "total_records": 0,
                "talk_minutes": 0,
                "billable_minutes": 0,
                "billed_amount": 0,
                "wallet_recharged": 0,
                "wallet_pending": 0,
                "wallet_remaining": 0,
            },
            "customer_summary": [],
            "records": [],
            "wallet": {"wallet_enabled": bool(get_wallet_transaction_columns())},
            "transactions": [],
        }
        load_error = str(exc)

    message, message_type = pop_ui_message()
    return render_template(
        "call_detail_record.html",
        payload=payload,
        load_error=load_error,
        message=message,
        message_type=message_type,
    )


@app.route("/call_detail_record_data")
def call_detail_record_data():
    if not login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    try:
        return jsonify({"ok": True, "payload": get_filtered_call_records(session["user"])})
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 502


@app.route("/wallet_recharge", methods=["POST"])
def wallet_recharge():
    if not login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    payload = request.get_json(silent=True)
    if payload is None:
        return jsonify({"ok": False, "error": "Invalid request body. Please refresh the page and try again."}), 400

    amount_raw = payload.get("amount")
    try:
        amount = float(amount_raw)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Please enter a valid recharge amount."}), 400

    if amount <= 0:
        return jsonify({"ok": False, "error": "Recharge amount must be greater than zero."}), 400

    if amount < MIN_RECHARGE_AMOUNT or amount > MAX_RECHARGE_AMOUNT:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": f"Recharge amount must be between Rs {MIN_RECHARGE_AMOUNT:.0f} and Rs {MAX_RECHARGE_AMOUNT:.0f}.",
                }
            ),
            400,
        )

    try:
        success, message, payment_reference = create_wallet_recharge_request(session["user"], amount)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Wallet transaction could not be created: {exc}"}), 500
    if not success:
        return jsonify({"ok": False, "error": message}), 400

    try:
        if razorpay_enabled():
            payment_link = create_razorpay_payment_link(session["user"], payment_reference, amount)
            update_wallet_gateway_link(
                payment_reference=payment_reference,
                gateway_link_id=payment_link["id"],
                gateway_link_url=payment_link["short_url"],
                gateway_payload=payment_link["payload"],
            )
            return (
                jsonify(
                    {
                        "ok": True,
                        "message": "Payment link created successfully.",
                        "payment_reference": payment_reference,
                        "redirect_url": payment_link["short_url"],
                    }
                ),
                200,
            )
    except Exception as exc:
        update_wallet_transaction_status(payment_reference, "failed", gateway_payload=str(exc))
        return jsonify({"ok": False, "error": f"Unable to create Razorpay payment link: {exc}"}), 502

    return (
        jsonify(
            {
                "ok": True,
                "message": "Razorpay API keys not configured. Using fallback recharge URL.",
                "payment_reference": payment_reference,
                "redirect_url": build_wallet_redirect_url(payment_reference, amount),
            }
        ),
        200,
    )


@app.route("/wallet/payment/callback", methods=["GET"])
def wallet_payment_callback():
    payment_reference = (request.args.get("razorpay_payment_link_reference_id") or "").strip()
    payment_link_id = (request.args.get("razorpay_payment_link_id") or "").strip()
    payment_status = (request.args.get("razorpay_payment_link_status") or "").strip().lower()
    payment_id = (request.args.get("razorpay_payment_id") or "").strip()
    signature = (request.args.get("razorpay_signature") or "").strip()

    transaction = get_wallet_transaction(payment_reference)
    if not transaction:
        set_ui_message("Payment callback received, but transaction reference was not found.", "error")
        return redirect("/call-detail-record")

    if not razorpay_enabled():
        set_ui_message("Razorpay keys are missing, so payment verification could not be completed.", "error")
        return redirect("/call-detail-record")

    if not verify_razorpay_callback_signature(
        payment_link_id=payment_link_id,
        payment_reference=payment_reference,
        payment_status=payment_status,
        payment_id=payment_id,
        signature=signature,
    ):
        update_wallet_transaction_status(
            payment_reference=payment_reference,
            payment_status="failed",
            gateway_payment_id=payment_id,
            gateway_payload=str(request.args.to_dict()),
        )
        set_ui_message("Payment signature verification failed. No wallet amount was added.", "error")
        return redirect("/call-detail-record")

    try:
        payment_link = fetch_razorpay_payment_link(payment_link_id)
    except Exception as exc:
        set_ui_message(f"Unable to confirm payment with Razorpay: {exc}", "error")
        return redirect("/call-detail-record")

    amount_paid = float(payment_link.get("amount_paid", 0) or 0) / 100
    verified_paid = (
        payment_link.get("status") == "paid"
        and payment_link.get("reference_id") == payment_reference
        and amount_paid >= transaction["amount"]
    )

    if verified_paid:
        update_wallet_transaction_status(
            payment_reference=payment_reference,
            payment_status="paid",
            gateway_payment_id=payment_id,
            gateway_order_id=payment_link.get("order_id", ""),
            gateway_signature=signature,
            gateway_payload=json.dumps(payment_link),
        )
        set_ui_message("Payment verified successfully with Razorpay and added to wallet.", "success")
        return redirect("/")

    update_wallet_transaction_status(
        payment_reference=payment_reference,
        payment_status="failed",
        gateway_payment_id=payment_id,
        gateway_signature=signature,
        gateway_payload=json.dumps(payment_link),
    )
    set_ui_message("Payment was not marked paid by Razorpay, so wallet was not credited.", "error")
    return redirect("/call-detail-record")


@app.route("/wallet/payment/success", methods=["GET", "POST"])
def wallet_payment_success():
    payment_reference = (request.values.get("payment_reference") or request.values.get("reference") or "").strip()
    gateway_payment_id = (request.values.get("razorpay_payment_id") or request.values.get("payment_id") or "").strip()
    gateway_order_id = (request.values.get("razorpay_order_id") or request.values.get("order_id") or "").strip()
    gateway_signature = (request.values.get("razorpay_signature") or request.values.get("signature") or "").strip()

    if not payment_reference:
        if login_required():
            set_ui_message("Payment success callback did not include a payment reference.", "error")
            return redirect("/")
        return jsonify({"ok": False, "error": "payment_reference is required."}), 400

    success, message = update_wallet_transaction_status(
        payment_reference=payment_reference,
        payment_status="paid",
        gateway_payment_id=gateway_payment_id,
        gateway_order_id=gateway_order_id,
        gateway_signature=gateway_signature,
        gateway_payload=str(request.values.to_dict()),
    )

    if request.method == "POST":
        return jsonify({"ok": success, "message": message}), 200 if success else 400

    set_ui_message(
        "Payment successful. Wallet balance has been updated." if success else message,
        "success" if success else "error",
    )
    return redirect("/")


@app.route("/wallet/payment/failed", methods=["GET", "POST"])
def wallet_payment_failed():
    payment_reference = (request.values.get("payment_reference") or request.values.get("reference") or "").strip()
    gateway_payment_id = (request.values.get("razorpay_payment_id") or request.values.get("payment_id") or "").strip()

    if not payment_reference:
        if login_required():
            set_ui_message("Payment was cancelled, but no payment reference was returned.", "error")
            return redirect("/call-detail-record")
        return jsonify({"ok": False, "error": "payment_reference is required."}), 400

    success, message = update_wallet_transaction_status(
        payment_reference=payment_reference,
        payment_status="failed",
        gateway_payment_id=gateway_payment_id,
        gateway_payload=str(request.values.to_dict()),
    )

    if request.method == "POST":
        return jsonify({"ok": success, "message": message}), 200 if success else 400

    set_ui_message(
        "Payment did not complete, so no amount was added to the wallet." if success else message,
        "error",
    )
    return redirect("/call-detail-record")


@app.route("/outbound", methods=["GET", "POST"])
def outbound():
    if not login_required():
        return redirect("/login")

    approved_agent = get_current_agent_profile(session["user"])

    if request.method == "POST":
        if not approved_agent:
            return redirect("/settings")

        phone = request.form["phone"].strip()
        schedule_time = parse_schedule_value(request.form["time"])
        schedule_single_call(session["user"], phone, schedule_time)
        return redirect("/outbound")

    message, message_type = pop_ui_message()
    return render_template(
        "outbound.html",
        approved_agent=approved_agent,
        message=message,
        message_type=message_type,
    )


@app.route("/schedule_call", methods=["POST"])
def schedule_call():
    approved_agent, error = require_approved_agent()
    if error:
        return error

    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    time_value = (data.get("time") or "").strip()

    if not phone or not time_value:
        return jsonify({"ok": False, "error": "Phone and time are required."}), 400

    try:
        schedule_time = parse_schedule_value(time_value)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400

    call_id = schedule_single_call(session["user"], phone, schedule_time)

    return jsonify(
        {
            "ok": True,
            "call_id": call_id,
            "status": "scheduled",
            "message": f"Call scheduled with approved agent {approved_agent['credential_id']}.",
        }
    )


@app.route("/upload_contacts", methods=["POST"])
def upload_contacts():
    _, error = require_approved_agent()
    if error:
        return error

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Please upload an Excel file."}), 400

    uploaded_file = request.files["file"]
    extension = os.path.splitext(uploaded_file.filename or "")[1].lower()

    if extension not in ALLOWED_UPLOAD_EXTENSIONS:
        return jsonify({"ok": False, "error": "Only .xlsx or .csv files are supported."}), 400

    default_time_raw = (request.form.get("default_time") or "").strip()
    default_schedule_time = None
    if default_time_raw:
        try:
            default_schedule_time = parse_schedule_value(default_time_raw)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

    try:
        if extension == ".csv":
            import csv
            import io

            stream = io.StringIO(uploaded_file.stream.read().decode("utf-8-sig"))
            reader = csv.DictReader(stream)
            parsed_rows = []
            for row in reader:
                normalized = {normalize_header(key): value for key, value in row.items()}
                phone = str(
                    next((normalized[key] for key in normalized if key in PHONE_HEADERS), "") or ""
                ).strip()
                if not phone:
                    continue

                time_value = next((normalized[key] for key in normalized if key in TIME_HEADERS), "")
                name_value = next((normalized[key] for key in normalized if key in NAME_HEADERS), "")
                schedule_time = parse_schedule_value(time_value) if time_value else default_schedule_time
                if schedule_time is None:
                    raise ValueError(f"Schedule time missing for phone {phone}.")

                parsed_rows.append(
                    {
                        "phone": phone,
                        "schedule_time": schedule_time,
                        "customer_name": str(name_value or "").strip(),
                    }
                )
        else:
            parsed_rows = parse_excel_rows(uploaded_file, default_schedule_time)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400

    if not parsed_rows:
        return jsonify({"ok": False, "error": "No valid contacts found in the uploaded file."}), 400

    scheduled_count = 0
    for row in parsed_rows:
        schedule_single_call(
            user_id=session["user"],
            phone=row["phone"],
            schedule_time=row["schedule_time"],
            contact_name=row.get("customer_name"),
        )
        scheduled_count += 1

    return jsonify(
        {
            "ok": True,
            "scheduled_count": scheduled_count,
            "message": f"{scheduled_count} calls uploaded and scheduled successfully.",
        }
    )


@app.route("/calls")
def calls():
    if not login_required():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    return jsonify(fetch_calls_for_user(session["user"]))


@app.route("/internal/call_status", methods=["POST"])
def internal_call_status():
    token = request.headers.get("X-Status-Token")
    if token != app.config["STATUS_UPDATE_TOKEN"]:
        return jsonify({"ok": False, "error": "Forbidden"}), 403

    data = request.get_json(silent=True) or {}
    call_id = data.get("call_id")
    status = (data.get("status") or "").strip()
    message = (data.get("message") or "").strip()

    if not call_id or not status:
        return jsonify({"ok": False, "error": "call_id and status are required."}), 400

    update_call_record(call_id, status, message or None)
    return jsonify({"ok": True})


@app.route("/settings", methods=["GET", "POST"])
def settings():
    if not login_required():
        return redirect("/login")

    message = ""
    message_type = ""

    if request.method == "POST":
        credential_id = request.form.get("credential_id", "").strip()
        credential_password = request.form.get("credential_password", "").strip()

        if not credential_id or not credential_password:
            message = "Please enter credential ID and password."
            message_type = "error"
        else:
            success, message = approve_agent_for_user(session["user"], credential_id, credential_password)
            message_type = "success" if success else "error"

    profile = get_current_agent_profile(session["user"])
    return render_template(
        "settings.html",
        profile=profile,
        message=message,
        message_type=message_type,
    )


@app.route("/transfer-settings", methods=["GET", "POST"])
def transfer_settings():
    if not login_required():
        return redirect("/login")

    message = ""
    message_type = ""

    if request.method == "POST":
        transfer_number = request.form.get("transfer_number", "").strip()
        is_enabled = bool(request.form.get("is_enabled"))

        if is_enabled and not transfer_number:
            message = "Please enter a transfer number before enabling live transfer."
            message_type = "error"
        else:
            success, message = save_transfer_settings(session["user"], transfer_number, is_enabled)
            message_type = "success" if success else "error"

    transfer_config = get_transfer_settings(session["user"])
    approved_agent = get_current_agent_profile(session["user"])
    return render_template(
        "transfer_settings.html",
        transfer_config=transfer_config,
        approved_agent=approved_agent,
        message=message,
        message_type=message_type,
    )


start_scheduler_once()


if __name__ == "__main__":
    start_scheduler_once()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
